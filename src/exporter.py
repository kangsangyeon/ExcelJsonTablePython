import ast
import json
import os
from pathlib import Path

import openpyxl

type_mappings = {
    'Integer': int,
    'String': str,
    'Float': float,
    'Boolean': bool,
    'Array': list,
    'Object': dict
}


def validate_and_convert_value(value, expected_type):
    """
    값이 기대한 타입인지 검증하고, 타입에 맞게 변환하는 함수.
    """
    try:
        if value is None:
            return None
        if expected_type == list:
            arr_str = '[' + str(value).replace('\n', ',') + ']'
            return ast.literal_eval(arr_str)
        if expected_type == dict:
            return ast.literal_eval(value)
        return expected_type(value)
    except (ValueError, TypeError) as e:
        raise ValueError(f"값 '{value}'는 기대한 타입 '{expected_type.__name__}'과 일치하지 않습니다: {e}")


def export_excel_to_json(excel_file_path):
    """
    엑셀 파일을 읽어 JSON 형식으로 저장하는 함수.

    :param excel_file_path: 엑셀 파일 경로
    """
    workbook = openpyxl.load_workbook(excel_file_path)

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]

        output_path = sheet['A1'].value
        # A1 셀이 비어있거나 .json으로 끝나지 않으면 이 시트를 제외
        if not output_path or not output_path.strip().lower().endswith('.json'):
            print(f"스킵: {excel_file_path}의 {sheet_name} 시트 - A1 셀이 비어있거나 .json으로 끝나지 않음")
            continue

        output_path = os.path.expandvars(output_path)
        
        # 절대경로가 아닌 경우 엑셀 파일의 디렉토리를 기준으로 경로 설정
        if not os.path.isabs(output_path):
            excel_dir = os.path.dirname(os.path.abspath(excel_file_path))
            output_path = os.path.join(excel_dir, output_path)

        output_dir = os.path.dirname(output_path)
        Path(output_dir).mkdir(parents=True, exist_ok=True)

        headers = []
        parsed_headers = []

        for cell in sheet[2]:
            if cell.value is None:
                break
            header_name, column_type_name = cell.value.split(":")
            header_name = header_name.strip()
            column_type = type_mappings.get(column_type_name.strip())
            headers.append(header_name)
            parsed_headers.append((header_name, column_type))

        data_list = []

        for row in sheet.iter_rows(min_row=3, values_only=True):
            if row[0] is None:
                break
            data = {}
            for value, (header_name, col_type) in zip(row, parsed_headers):
                converted_value = validate_and_convert_value(value, col_type)
                if converted_value is not None:
                    data[header_name] = converted_value

            data_list.append(data)

        with open(output_path, 'w', encoding='utf-8') as json_file:
            json.dump(data_list, json_file, ensure_ascii=False, indent=2)

        print(f"JSON 파일이 성공적으로 저장되었습니다: {output_path}")


def export_excel_to_json_dir(directory):
    """
    지정된 디렉토리 내 모든 Excel 파일을 처리하는 함수.

    :param directory: Excel 파일들이 위치한 디렉토리 경로
    """
    for filename in os.listdir(directory):
        if filename.startswith('~$'):
            continue
        if filename.endswith(('.xlsx', '.xlsm', '.xltx')):
            file_path = os.path.join(directory, filename)
            try:
                export_excel_to_json(file_path)
            except Exception as e:
                print(f"파일 처리 중 오류 발생: {filename} - {e}")
